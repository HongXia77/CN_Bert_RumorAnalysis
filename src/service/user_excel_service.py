from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel as from_excel_serial
from sqlalchemy import or_
from sqlalchemy.orm import Session

from modules.tools.logger import logger
from src.models._init_database import User
from src.utils.security import get_password_hash


EXPORT_HEADERS = [
    ("user_id", "用户ID"),
    ("username", "用户名"),
    ("password", "密码"),
    ("email", "邮箱"),
    ("phone", "手机号"),
    ("role", "角色"),
    ("gender", "性别"),
    ("province", "省份"),
    ("city", "城市/区县"),
    ("birthday", "生日"),
    ("status", "状态"),
    ("avatar", "头像"),
    ("create_time", "创建时间"),
    ("update_time", "更新时间"),
]

IMPORT_HEADER_ALIASES = {
    "user_id": {"userid", "userid", "用户id", "用户编号", "用户id号"},
    "username": {"username", "用户名", "账号", "登录名"},
    "password": {"password", "密码"},
    "email": {"email", "邮箱", "电子邮箱"},
    "phone": {"phone", "手机号", "手机号码", "电话"},
    "role": {"role", "角色"},
    "gender": {"gender", "性别"},
    "province": {"province", "省份"},
    "city": {"city", "城市", "区县", "城市/区县"},
    "birthday": {"birthday", "生日", "出生日期"},
    "status": {"status", "状态"},
    "avatar": {"avatar", "头像", "头像地址"},
}

ROLE_LABEL_MAP = {
    "admin": "admin",
    "user": "user",
    "管理员": "admin",
    "普通用户": "user",
}

STATUS_LABEL_MAP = {
    "正常": "正常",
    "禁用": "禁用",
    "未激活": "未激活",
}

GENDER_LABEL_MAP = {
    "男": "男",
    "女": "女",
    "未知": "未知",
}

MAX_ERROR_DETAILS = 20


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("_", "").replace(" ", "")


HEADER_FIELD_MAP = {
    _normalize_header(alias): field
    for field, aliases in IMPORT_HEADER_ALIASES.items()
    for alias in aliases
}


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_optional_int(value: Any) -> int | None:
    text = _normalize_text(value)
    if not text:
        return None
    if text.endswith(".0"):
        text = text[:-2]
    return int(text)


def _parse_birthday(value: Any) -> date | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        numeric = float(value)
        if numeric > 10_000_000_000:
            return datetime.fromtimestamp(numeric / 1000).date()
        if numeric > 1_000_000_000:
            return datetime.fromtimestamp(numeric).date()
        if 20_000 <= numeric <= 80_000:
            excel_value = from_excel_serial(numeric)
            return excel_value.date() if isinstance(excel_value, datetime) else excel_value

    text = _normalize_text(value)
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    raise ValueError("生日格式不正确，支持 YYYY-MM-DD / YYYY/MM/DD")


def _serialize_datetime(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""


def _serialize_date(value: date | None) -> str:
    return value.isoformat() if value else ""


def _build_export_row(user: User) -> list[str | int]:
    return [
        user.user_id,
        user.username,
        "",
        user.email,
        user.phone,
        user.role,
        user.gender,
        user.province or "",
        user.city or "",
        _serialize_date(user.birthday),
        user.status,
        user.avatar or "",
        _serialize_datetime(user.create_time),
        _serialize_datetime(user.update_time),
    ]


def _build_workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    buffer.seek(0)
    return buffer.getvalue()


def build_user_import_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "用户导入模板"
    worksheet.append([label for _, label in EXPORT_HEADERS])

    note_sheet = workbook.create_sheet("填写说明")
    note_sheet.append(["说明"])
    note_sheet.append(["1. 支持导入 .xlsx 文件，默认读取第一个工作表。"])
    note_sheet.append(["2. user_id 为空时视为新增用户；填写已存在 user_id 时视为更新该用户。"])
    note_sheet.append(["3. 新增用户必须填写：用户名、密码、邮箱、手机号。"])
    note_sheet.append(["4. 更新用户时，密码可留空；留空表示不修改密码。"])
    note_sheet.append(["5. 角色支持：admin / user / 管理员 / 普通用户。"])
    note_sheet.append(["6. 生日支持 YYYY-MM-DD、YYYY/MM/DD 或 Excel 日期格式。"])

    return _build_workbook_bytes(workbook)


def export_users_to_excel(db: Session, *, search: str | None = None, role: str | None = None, status: str | None = None) -> bytes:
    query = db.query(User)
    if search:
        keyword = f"%{search.strip()}%"
        query = query.filter(
            or_(
                User.username.like(keyword),
                User.email.like(keyword),
                User.phone.like(keyword),
            )
        )
    if role:
        query = query.filter(User.role == role)
    if status:
        query = query.filter(User.status == status)

    users = query.order_by(User.create_time.desc(), User.user_id.desc()).all()

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("用户数据")
    worksheet.append([label for _, label in EXPORT_HEADERS])

    for user in users:
        worksheet.append(_build_export_row(user))

    return _build_workbook_bytes(workbook)


def _resolve_header_mapping(header_row: tuple[Any, ...] | None) -> dict[str, int]:
    if not header_row:
        raise ValueError("Excel 文件缺少表头")

    mapping: dict[str, int] = {}
    for index, cell_value in enumerate(header_row):
        field_name = HEADER_FIELD_MAP.get(_normalize_header(cell_value))
        if field_name and field_name not in mapping:
            mapping[field_name] = index

    if "username" not in mapping and "user_id" not in mapping:
        raise ValueError("导入表头缺少用户标识，至少需要 用户ID 或 用户名")

    return mapping


def _read_row_value(row: tuple[Any, ...], header_mapping: dict[str, int], field_name: str) -> Any:
    index = header_mapping.get(field_name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _normalize_role(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return "user"
    role = ROLE_LABEL_MAP.get(text)
    if role is None:
        raise ValueError("角色仅支持 admin / user / 管理员 / 普通用户")
    return role


def _normalize_status(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return "正常"
    status = STATUS_LABEL_MAP.get(text)
    if status is None:
        raise ValueError("状态仅支持 正常 / 禁用 / 未激活")
    return status


def _normalize_gender(value: Any) -> str:
    text = _normalize_text(value)
    if not text:
        return "未知"
    gender = GENDER_LABEL_MAP.get(text)
    if gender is None:
        raise ValueError("性别仅支持 男 / 女 / 未知")
    return gender


def _parse_import_rows(file_bytes: bytes) -> tuple[list[dict[str, Any]], int]:
    workbook = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        header_mapping = _resolve_header_mapping(header_row)

        parsed_rows: list[dict[str, Any]] = []
        blank_rows = 0

        for row_number, row in enumerate(rows, start=2):
            raw_values = {
                field: _read_row_value(row, header_mapping, field)
                for field in header_mapping
            }
            if all(_normalize_text(value) == "" for value in raw_values.values()):
                blank_rows += 1
                continue

            parsed_rows.append(
                {
                    "row_number": row_number,
                    "user_id": _parse_optional_int(raw_values.get("user_id")),
                    "username": _normalize_text(raw_values.get("username")),
                    "password": _normalize_text(raw_values.get("password")),
                    "email": _normalize_text(raw_values.get("email")),
                    "phone": _normalize_text(raw_values.get("phone")),
                    "role": _normalize_role(raw_values.get("role")),
                    "gender": _normalize_gender(raw_values.get("gender")),
                    "province": _normalize_text(raw_values.get("province")),
                    "city": _normalize_text(raw_values.get("city")),
                    "birthday": _parse_birthday(raw_values.get("birthday")),
                    "status": _normalize_status(raw_values.get("status")),
                    "avatar": _normalize_text(raw_values.get("avatar")),
                }
            )
    finally:
        workbook.close()

    return parsed_rows, blank_rows


def _append_error(errors: list[str], message: str) -> None:
    if len(errors) < MAX_ERROR_DETAILS:
        errors.append(message)


def _find_target_user(
    row: dict[str, Any],
    *,
    users_by_id: dict[int, User],
    users_by_username: dict[str, User],
    users_by_email: dict[str, User],
    users_by_phone: dict[str, User],
) -> User | None:
    matched_users: set[User] = set()

    if row["user_id"] is not None:
        user = users_by_id.get(row["user_id"])
        if user is None:
            raise ValueError(f"用户ID {row['user_id']} 不存在")
        matched_users.add(user)

    for field_name, mapping in (
        ("username", users_by_username),
        ("email", users_by_email),
        ("phone", users_by_phone),
    ):
        value = row[field_name]
        if value and value in mapping:
            matched_users.add(mapping[value])

    if len(matched_users) > 1:
        raise ValueError("同一行匹配到了多个不同用户，请检查 用户ID / 用户名 / 邮箱 / 手机号")

    return next(iter(matched_users), None)


def import_users_from_excel(db: Session, *, current_user: User, file_bytes: bytes) -> tuple[dict[str, Any] | None, dict[str, Any] | None]:
    try:
        parsed_rows, blank_rows = _parse_import_rows(file_bytes)
    except ValueError as exc:
        return None, {"message": str(exc), "errors": [str(exc)]}
    except Exception as exc:
        logger.error(f"解析用户导入 Excel 失败: {exc}", exc_info=True)
        return None, {"message": "Excel 解析失败，请检查文件格式", "errors": ["文件格式异常或内容损坏"]}

    if not parsed_rows:
        return None, {"message": "Excel 中没有可导入的数据", "errors": ["未检测到有效数据行"]}

    existing_users = db.query(User).order_by(User.user_id.asc()).all()
    users_by_id = {user.user_id: user for user in existing_users}
    users_by_username = {user.username: user for user in existing_users}
    users_by_email = {user.email: user for user in existing_users}
    users_by_phone = {user.phone: user for user in existing_users}

    errors: list[str] = []
    processed_user_ids: set[int] = set()
    created_count = 0
    updated_count = 0
    new_users: list[User] = []

    try:
        for row in parsed_rows:
            try:
                target_user = _find_target_user(
                    row,
                    users_by_id=users_by_id,
                    users_by_username=users_by_username,
                    users_by_email=users_by_email,
                    users_by_phone=users_by_phone,
                )
            except ValueError as exc:
                _append_error(errors, f"第 {row['row_number']} 行：{exc}")
                continue

            if target_user is not None:
                if target_user.user_id in processed_user_ids:
                    _append_error(errors, f"第 {row['row_number']} 行：同一个用户在文件中出现了多次")
                    continue

                next_username = row["username"] or target_user.username
                next_email = row["email"] or target_user.email
                next_phone = row["phone"] or target_user.phone

                for field_name, next_value, mapping in (
                    ("用户名", next_username, users_by_username),
                    ("邮箱", next_email, users_by_email),
                    ("手机号", next_phone, users_by_phone),
                ):
                    conflict_user = mapping.get(next_value)
                    if conflict_user is not None and conflict_user.user_id != target_user.user_id:
                        _append_error(errors, f"第 {row['row_number']} 行：{field_name} {next_value} 已被其他用户占用")
                        break
                else:
                    if target_user.user_id == current_user.user_id and row["role"] != "admin":
                        _append_error(errors, f"第 {row['row_number']} 行：不能通过导入取消当前管理员自己的权限")
                        continue
                    if target_user.user_id == current_user.user_id and row["status"] != "正常":
                        _append_error(errors, f"第 {row['row_number']} 行：不能通过导入将当前管理员状态改为不可用")
                        continue

                    users_by_username.pop(target_user.username, None)
                    users_by_email.pop(target_user.email, None)
                    users_by_phone.pop(target_user.phone, None)

                    target_user.username = next_username
                    target_user.email = next_email
                    target_user.phone = next_phone
                    target_user.role = row["role"]
                    target_user.gender = row["gender"]
                    target_user.province = row["province"]
                    target_user.city = row["city"]
                    target_user.birthday = row["birthday"]
                    target_user.status = row["status"]
                    target_user.avatar = row["avatar"]
                    if row["password"]:
                        target_user.password = get_password_hash(row["password"])

                    users_by_username[target_user.username] = target_user
                    users_by_email[target_user.email] = target_user
                    users_by_phone[target_user.phone] = target_user
                    processed_user_ids.add(target_user.user_id)
                    updated_count += 1
                    continue

                continue

            missing_fields = [field for field in ("username", "password", "email", "phone") if not row[field]]
            if missing_fields:
                missing_labels = {
                    "username": "用户名",
                    "password": "密码",
                    "email": "邮箱",
                    "phone": "手机号",
                }
                _append_error(
                    errors,
                    f"第 {row['row_number']} 行：新增用户必须填写 {'、'.join(missing_labels[field] for field in missing_fields)}",
                )
                continue

            for field_name, next_value, mapping in (
                ("用户名", row["username"], users_by_username),
                ("邮箱", row["email"], users_by_email),
                ("手机号", row["phone"], users_by_phone),
            ):
                if next_value in mapping:
                    _append_error(errors, f"第 {row['row_number']} 行：{field_name} {next_value} 已存在")
                    break
            else:
                new_user = User(
                    username=row["username"],
                    password=get_password_hash(row["password"]),
                    email=row["email"],
                    phone=row["phone"],
                    avatar=row["avatar"],
                    role=row["role"],
                    gender=row["gender"],
                    province=row["province"],
                    city=row["city"],
                    birthday=row["birthday"],
                    status=row["status"],
                )
                new_users.append(new_user)
                users_by_username[new_user.username] = new_user
                users_by_email[new_user.email] = new_user
                users_by_phone[new_user.phone] = new_user
                created_count += 1

        if errors:
            db.rollback()
            return None, {
                "message": "导入校验未通过，请修正后重试",
                "errors": errors,
                "summary": {
                    "total_rows": len(parsed_rows),
                    "blank_rows": blank_rows,
                    "created": created_count,
                    "updated": updated_count,
                    "failed": len(errors),
                },
            }

        future_admin_count = sum(1 for user in existing_users if user.role == "admin") + sum(
            1 for user in new_users if user.role == "admin"
        )
        if future_admin_count <= 0:
            db.rollback()
            return None, {
                "message": "导入后系统至少需要保留一名管理员",
                "errors": ["请至少保留一名管理员账号"],
            }

        if new_users:
            db.add_all(new_users)

        db.commit()
        logger.business(
            "管理员批量导入用户成功 => "
            f"admin_id={current_user.user_id}, total_rows={len(parsed_rows)}, created={created_count}, "
            f"updated={updated_count}, blank_rows={blank_rows}"
        )
        return {
            "total_rows": len(parsed_rows),
            "blank_rows": blank_rows,
            "created": created_count,
            "updated": updated_count,
            "failed": 0,
        }, None
    except Exception as exc:
        db.rollback()
        logger.error(f"管理员批量导入用户失败: {exc}", exc_info=True)
        return None, {
            "message": "批量导入失败，请稍后重试",
            "errors": ["服务端处理导入文件时出现异常"],
        }
